import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

NAVY = "1E2A5E"
GREEN = "00A878"
WHITE = "FFFFFF"
GREY = "F2F1ED"

FONT = "Arial"
h_font = Font(name=FONT, bold=True, color=WHITE, size=11)
h_fill = PatternFill("solid", start_color=NAVY)
title_font = Font(name=FONT, bold=True, size=14, color=NAVY)
base_font = Font(name=FONT, size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

monthly = pd.read_csv("/tmp/_monthly.csv")
regional = pd.read_csv("/tmp/_regional.csv")
segmento = pd.read_csv("/tmp/_segmento.csv")
setor = pd.read_csv("/tmp/_setor.csv")
funil = pd.read_csv("/tmp/_funil.csv")
risco = pd.read_csv("/tmp/_risco.csv")
ops = pd.read_csv("/tmp/_ops_diario.csv")

wb = Workbook()


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def write_df(ws, df, start_row=3, start_col=1, number_formats=None):
    number_formats = number_formats or {}
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    style_header(ws, start_row, len(df.columns))
    for i, row_vals in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row_vals):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=val)
            cell.font = base_font
            cell.border = border
            colname = df.columns[j]
            if colname in number_formats:
                cell.number_format = number_formats[colname]
    return start_row + len(df) + 1  # next free row


def autosize(ws, ncols, width=16):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


# ---------------- Sheet: Capa ----------------
ws = wb.active
ws.title = "Capa"
ws["B2"] = "Ippiz — Base de Dados de Carteira e Risco"
ws["B2"].font = Font(name=FONT, bold=True, size=18, color=NAVY)
ws["B4"] = "Crédito instantâneo via PIX para micro e pequenas empresas (dados 100% fictícios)"
ws["B4"].font = Font(name=FONT, size=11, italic=True)
ws["B6"] = "Case de portfólio — construído a partir da descrição de vaga de Business Analyst"
ws["B6"].font = base_font
ws["B8"] = "Abas:"
ws["B8"].font = Font(name=FONT, bold=True)
abas = [
    "Serie_Mensal — evolução de carteira, inadimplência, SLA, aprovação (18 meses)",
    "Regional / Segmento / Setor — quebras da carteira ativa atual",
    "Funil_Aprovacao — solicitado > análise > aprovado > desembolsado > recusado",
    "Faixas_Risco — aging da carteira (em dia até 90+ dias)",
    "Operacao_Diaria — produtividade e fila por analista (últimos 30 dias)",
    "KPIs — indicadores consolidados do mês corrente, com fórmulas",
]
for i, a in enumerate(abas):
    ws.cell(row=9 + i, column=2, value=f"•  {a}").font = base_font
ws.column_dimensions["B"].width = 100
ws.sheet_view.showGridLines = False

# ---------------- Sheet: Serie_Mensal ----------------
ws = wb.create_sheet("Serie_Mensal")
ws["A1"] = "Série Mensal — Ippiz"
ws["A1"].font = title_font
nf = {
    "valor_concedido": '#,##0;(#,##0);"-"',
    "ticket_medio": '#,##0;(#,##0);"-"',
    "inadimplencia_15d": '0.0"%"',
    "inadimplencia_30d": '0.0"%"',
    "inadimplencia_90d": '0.0"%"',
    "taxa_aprovacao": '0.0"%"',
    "sla_medio_horas": '0.0',
    "recuperacao_cobranca": '0.0"%"',
    "nps": '0.0',
}
next_row = write_df(ws, monthly, start_row=3, number_formats=nf)
autosize(ws, len(monthly.columns), 15)

# gráfico de evolução de carteira
chart = LineChart()
chart.title = "Valor Concedido por Mês (R$)"
chart.style = 2
data_ref = Reference(ws, min_col=3, max_col=3, min_row=3, max_row=3 + len(monthly))
cats_ref = Reference(ws, min_col=2, min_row=4, max_row=3 + len(monthly))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width, chart.height = 24, 10
ws.add_chart(chart, f"A{next_row + 2}")

chart2 = LineChart()
chart2.title = "Inadimplência 30d (%) e Taxa de Aprovação (%)"
d2 = Reference(ws, min_col=7, max_col=7, min_row=3, max_row=3 + len(monthly))
d3 = Reference(ws, min_col=9, max_col=9, min_row=3, max_row=3 + len(monthly))
chart2.add_data(d2, titles_from_data=True)
chart2.add_data(d3, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.width, chart2.height = 24, 10
ws.add_chart(chart2, f"A{next_row + 22}")

# ---------------- Sheet: Regional ----------------
ws = wb.create_sheet("Regional")
ws["A1"] = "Carteira Ativa por Região"
ws["A1"].font = title_font
nfr = {"valor_carteira": '#,##0;(#,##0);"-"', "qtd_clientes_ativos": '#,##0', "inadimplencia_30d": '0.0"%"', "ticket_medio": '#,##0'}
nr = write_df(ws, regional, start_row=3, number_formats=nfr)
autosize(ws, len(regional.columns), 20)
ws[f"A{nr + 1}"] = "% da Carteira Total"
ws[f"A{nr + 1}"].font = Font(name=FONT, bold=True)
total_row = 3 + len(regional) + 1
for i in range(len(regional)):
    r = 4 + i
    ws.cell(row=nr + 2 + i, column=1, value=f"=B{r}/SUM(B4:B{3+len(regional)})")
    ws.cell(row=nr + 2 + i, column=1).number_format = "0.0%"

bchart = BarChart()
bchart.title = "Carteira Ativa por Região (R$)"
d = Reference(ws, min_col=2, max_col=2, min_row=3, max_row=3 + len(regional))
cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(regional))
bchart.add_data(d, titles_from_data=True)
bchart.set_categories(cats)
bchart.width, bchart.height = 20, 10
ws.add_chart(bchart, f"A{nr + 2 + len(regional) + 2}")

# ---------------- Sheet: Segmento ----------------
ws = wb.create_sheet("Segmento")
ws["A1"] = "Carteira Ativa por Segmento"
ws["A1"].font = title_font
write_df(ws, segmento, start_row=3, number_formats=nfr)
autosize(ws, len(segmento.columns), 24)

# ---------------- Sheet: Setor ----------------
ws = wb.create_sheet("Setor")
ws["A1"] = "Carteira Ativa por Setor"
ws["A1"].font = title_font
write_df(ws, setor, start_row=3, number_formats={"valor_carteira": '#,##0;(#,##0);"-"', "qtd_clientes_ativos": '#,##0'})
autosize(ws, len(setor.columns), 20)

# ---------------- Sheet: Funil_Aprovacao ----------------
ws = wb.create_sheet("Funil_Aprovacao")
ws["A1"] = "Funil de Aprovação — Mês Corrente"
ws["A1"].font = title_font
nr = write_df(ws, funil, start_row=3, number_formats={"qtd": "#,##0"})
autosize(ws, len(funil.columns), 16)
ws[f"E3"] = "% Conversão vs. Solicitado"
ws["E3"].font = h_font
ws["E3"].fill = h_fill
for i in range(len(funil)):
    r = 4 + i
    ws.cell(row=r, column=5, value=f"=C{r}/$C$4")
    ws.cell(row=r, column=5).number_format = "0.0%"
    ws.cell(row=r, column=5).border = border

# ---------------- Sheet: Faixas_Risco ----------------
ws = wb.create_sheet("Faixas_Risco")
ws["A1"] = "Faixas de Risco (Aging da Carteira)"
ws["A1"].font = title_font
write_df(ws, risco, start_row=3, number_formats={"valor": '#,##0;(#,##0);"-"', "percentual": '0.00"%"'})
autosize(ws, len(risco.columns), 16)

# ---------------- Sheet: Operacao_Diaria ----------------
ws = wb.create_sheet("Operacao_Diaria")
ws["A1"] = "Operação Diária por Analista (últimos 30 dias)"
ws["A1"].font = title_font
write_df(ws, ops, start_row=3, number_formats={"tempo_medio_resposta_h": "0.0"})
autosize(ws, len(ops.columns), 20)

# ---------------- Sheet: KPIs ----------------
ws = wb.create_sheet("KPIs")
ws["A1"] = "Indicadores Consolidados — Mês Corrente"
ws["A1"].font = title_font
last_row_monthly = 3 + len(monthly)
kpi_defs = [
    ("Valor Concedido no Mês (R$)", f"=Serie_Mensal!C{last_row_monthly}", '#,##0'),
    ("Variação vs. Mês Anterior (%)", f"=Serie_Mensal!C{last_row_monthly}/Serie_Mensal!C{last_row_monthly-1}-1", "0.0%"),
    ("Qtd. Operações no Mês", f"=Serie_Mensal!D{last_row_monthly}", "#,##0"),
    ("Ticket Médio (R$)", f"=Serie_Mensal!E{last_row_monthly}", "#,##0"),
    ("Taxa de Aprovação (%)", f"=Serie_Mensal!I{last_row_monthly}", "0.0%"),
    ("Inadimplência 15d (%)", f"=Serie_Mensal!F{last_row_monthly}", "0.0%"),
    ("Inadimplência 30d (%)", f"=Serie_Mensal!G{last_row_monthly}", "0.0%"),
    ("Inadimplência 90d (%)", f"=Serie_Mensal!H{last_row_monthly}", "0.0%"),
    ("SLA Médio de Análise (horas)", f"=Serie_Mensal!J{last_row_monthly}", "0.0"),
    ("Recuperação em Cobrança (%)", f"=Serie_Mensal!K{last_row_monthly}", "0.0%"),
    ("NPS", f"=Serie_Mensal!L{last_row_monthly}", "0.0"),
    ("Carteira Ativa Total (R$)", "=SUM(Regional!B4:B8)", "#,##0"),
    ("Clientes Ativos (Total)", "=SUM(Regional!C4:C8)", "#,##0"),
]
ws.cell(row=3, column=1, value="Indicador").font = h_font
ws.cell(row=3, column=1).fill = h_fill
ws.cell(row=3, column=2, value="Valor").font = h_font
ws.cell(row=3, column=2).fill = h_fill
for i, (label, formula, fmt) in enumerate(kpi_defs):
    r = 4 + i
    ws.cell(row=r, column=1, value=label).font = base_font
    c = ws.cell(row=r, column=2, value=formula)
    c.font = Font(name=FONT, bold=True, color=NAVY)
    c.number_format = fmt
    ws.cell(row=r, column=1).border = border
    c.border = border
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 20

wb.save("/home/claude/ippiz-portfolio/excel/Ippiz_Base_Dados.xlsx")
print("xlsx salvo")
